#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SOURCE_WORKBOOK = SCRIPT_DIR / "不山电商日报_2026年06.xlsx"
TARGET_WORKBOOK = SCRIPT_DIR / "不山电商日报_2026年07.xlsx"

DATE_SHEETS = [
    "每日明细+全平台汇总",
    "分平台销售占比",
    "渠道汇总",
    "全平台GMV日报",
    "平台GMV退款汇总",
]

MONTH_SHEET_OLD = "6月平台GMV完成情况"
MONTH_SHEET_NEW = "7月平台GMV完成情况"

MONTH_ROWS = [
    ("小红书直播", 600000, "小红书直播累计至7/1"),
    ("抖音直播", 650000, "抖音直播累计至7/1"),
    ("商品卡", 450000, "小红书商品卡+抖音商品卡累计至7/1"),
    ("视频号大号直播", 1250000, "视频号大号直播累计至7/1"),
    ("视频号小号直播", 200000, "视频号小号直播累计至7/1"),
    ("有赞", 1250000, "有赞累计至7/1"),
]


def clear_sheet_rows(ws, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def main() -> int:
    if TARGET_WORKBOOK.exists():
        print(f"已存在 7月 workbook：{TARGET_WORKBOOK}")
        return 0
    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(f"缺少 6月模板 workbook：{SOURCE_WORKBOOK}")

    wb = load_workbook(SOURCE_WORKBOOK)

    for sheet_name in DATE_SHEETS:
        if sheet_name in wb.sheetnames:
            clear_sheet_rows(wb[sheet_name], 2)

    if MONTH_SHEET_OLD in wb.sheetnames:
        ws = wb[MONTH_SHEET_OLD]
        ws.title = MONTH_SHEET_NEW
    elif MONTH_SHEET_NEW in wb.sheetnames:
        ws = wb[MONTH_SHEET_NEW]
    else:
        raise RuntimeError("未找到月度完成情况 sheet")

    for row_idx, (label, goal, desc) in enumerate(MONTH_ROWS, start=2):
        ws.cell(row_idx, 1).value = label
        ws.cell(row_idx, 2).value = 0
        ws.cell(row_idx, 3).value = 0
        ws.cell(row_idx, 4).value = 0
        ws.cell(row_idx, 5).value = goal
        ws.cell(row_idx, 6).value = 0
        ws.cell(row_idx, 7).value = 0
        ws.cell(row_idx, 8).value = desc
        ws.cell(row_idx, 9).value = None

    ws.cell(8, 1).value = "月度合计"
    ws.cell(8, 2).value = 0
    ws.cell(8, 3).value = 0
    ws.cell(8, 4).value = 0
    ws.cell(8, 5).value = 4400000
    ws.cell(8, 6).value = 0
    ws.cell(8, 7).value = "100.00%"
    ws.cell(8, 8).value = "完成度按实际成交额/月度目标计算。"
    ws.cell(8, 9).value = None
    ws.cell(9, 1).value = "说明：完成度按实际成交额/月度目标计算。数据已更新至7月1日；月度目标合计440.0万。"
    for col in range(2, ws.max_column + 1):
        ws.cell(9, col).value = None

    wb.save(TARGET_WORKBOOK)
    print(f"已创建 7月 workbook：{TARGET_WORKBOOK}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
