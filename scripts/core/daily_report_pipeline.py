#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
import subprocess
import shutil
import sys
import urllib.request
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


BASE_DIR = Path("/Users/jixukuangdian/Desktop/执行总监数据复盘")
DEFAULT_CONFIG = BASE_DIR / "daily_report_pipeline_config.json"
DEFAULT_WORKBOOK = BASE_DIR / "不山电商日报.xlsx"
DEFAULT_INPUT_DIR = BASE_DIR / "daily_report_inputs"
DEFAULT_OUTPUT_DIR = BASE_DIR / "daily_report_outputs"
REPORT_INSERT_ROW = 33
WEBHOOK_FALLBACK = "https://open.feishu.cn/open-apis/bot/v2/hook/7b4b324c-aaa4-486e-b946-90a9a83a7de6"

SPLIT_COLUMN_MAP = {
    ("小红书", "直播"): "小红书直播",
    ("小红书", "商品卡"): "小红书商品卡",
    ("有赞", "商城"): "有赞",
    ("抖音", "直播间"): "抖音直播间",
    ("视频号", "大号"): "视频号大号",
    ("视频号", "小号"): "视频号小号",
    ("抖音", "商品卡"): "抖音商品卡",
}

PLATFORM_ORDER = [
    ("小红书", "直播"),
    ("小红书", "商品卡"),
    ("抖音", "直播间"),
    ("抖音", "商品卡"),
    ("有赞", "商城"),
    ("视频号", "大号"),
    ("视频号", "小号"),
]

BASE_SYNC_TABLES = {
    "detail": "每日明细+全平台汇总",
    "channel": "渠道汇总",
    "share": "分平台销售占比",
    "april_progress": "平台GMV完成情况",
    "gmv_report": "全平台GMV日报",
}

MONTH_PROGRESS_ITEMS = [
    {
        "snapshot_key": "小红书直播",
        "target_key": "小红书直播",
        "summary_label": "小红书直播",
        "sheet_label": "小红书直播",
        "remark_label": "小红书直播",
    },
    {
        "snapshot_key": "抖音直播",
        "target_key": "抖音直播间",
        "summary_label": "抖音直播",
        "sheet_label": "抖音直播",
        "remark_label": "抖音直播",
    },
    {
        "snapshot_key": "商品卡",
        "target_key": "商品卡",
        "summary_label": "商品卡",
        "sheet_label": "商品卡",
        "remark_label": "小红书商品卡+抖音商品卡+视频号商品卡",
    },
    {
        "snapshot_key": "视频号大号",
        "target_key": "视频号大号",
        "summary_label": "视频号大号",
        "sheet_label": "视频号大号直播",
        "remark_label": "视频号大号直播",
    },
    {
        "snapshot_key": "视频号小号",
        "target_key": "视频号小号",
        "summary_label": "视频号小号",
        "sheet_label": "视频号小号直播",
        "remark_label": "视频号小号直播",
    },
    {
        "snapshot_key": "有赞",
        "target_key": "有赞",
        "summary_label": "有赞",
        "sheet_label": "有赞",
        "remark_label": "有赞",
    },
]


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def drop_none_values(payload: dict) -> dict:
    return {key: value for key, value in payload.items() if value is not None}


def round2(value: float) -> float:
    return round(float(value), 2)


def round6(value: float) -> float:
    return round(float(value), 6)


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def excel_date_to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:10]).date()
        except ValueError:
            return None
    return None


def format_wan(value: float) -> str:
    return f"{value / 10000:.2f}万"


def format_pct(value: float) -> str:
    return f"{value:.2%}"


def safe_div(numerator: float, denominator: float) -> float:
    if not denominator:
        return 0.0
    return numerator / denominator


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def normalize_entry(entry: dict) -> dict:
    platform = str(entry["platform"]).strip()
    account = str(entry["account"]).strip()
    gmv = round2(entry["gmv"])
    refund = round2(entry.get("refund", 0))
    actual = entry.get("actual")
    if actual is None:
        actual = round2(gmv - refund)
    else:
        actual = round2(actual)
    refund_rate = entry.get("refund_rate")
    if refund_rate is None:
        refund_rate = safe_div(refund, gmv)
    orders = entry.get("orders")
    buyers = entry.get("buyers")
    status = str(entry.get("status", "完整")).strip() or "完整"
    source = str(entry.get("source", "")).strip()
    report_label = str(entry.get("report_label", account)).strip() or account
    return {
        "platform": platform,
        "account": account,
        "report_label": report_label,
        "gmv": gmv,
        "actual": actual,
        "refund": refund,
        "refund_rate": refund_rate,
        "orders": orders,
        "buyers": buyers,
        "status": status,
        "source": source,
        "live_actual": round2(entry.get("live_actual", actual if platform == "视频号" else 0) or 0) if platform == "视频号" else round2(entry.get("live_actual", 0) or 0),
        "card_actual": round2(entry.get("card_actual", 0) or 0),
    }


def sort_entries(entries: list[dict]) -> list[dict]:
    order = {key: index for index, key in enumerate(PLATFORM_ORDER)}
    return sorted(entries, key=lambda item: (order.get((item["platform"], item["account"]), 999), item["platform"], item["account"]))


def validate_input(payload: dict) -> dict:
    report_date = parse_date(payload["date"])
    raw_entries = payload.get("entries", [])
    if not raw_entries:
        raise ValueError("输入文件缺少 entries，至少要有一个平台明细")
    entries = sort_entries([normalize_entry(entry) for entry in raw_entries])
    return {
        "date": report_date,
        "coverage_note": str(payload.get("coverage_note", "")).strip(),
        "conclusion": str(payload.get("conclusion", "")).strip(),
        "entries": entries,
        "totals_override": payload.get("totals_override") or {},
        "compare_override": payload.get("compare_override") or {},
        "month_progress_override": payload.get("month_progress_override") or {},
        "youzan_month_override": payload.get("youzan_month_override") or {},
    }


def load_config(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"未找到配置文件：{path}")
    return load_json(path)


def workbook_has_date(ws, target_date: date) -> bool:
    for row in range(2, ws.max_row + 1):
        if excel_date_to_date(ws.cell(row, 1).value) == target_date:
            return True
    return False


def find_rows_by_date(ws, target_date: date) -> list[int]:
    rows = []
    for row in range(2, ws.max_row + 1):
        if excel_date_to_date(ws.cell(row, 1).value) == target_date:
            rows.append(row)
    return rows


def delete_rows_by_date(ws, target_date: date) -> int:
    rows = find_rows_by_date(ws, target_date)
    for row in reversed(rows):
        ws.delete_rows(row, 1)
    return len(rows)


def find_insert_row_by_date(ws, target_date: date, date_col: int = 1) -> int:
    for row in range(2, ws.max_row + 1):
        row_date = excel_date_to_date(ws.cell(row, date_col).value)
        if row_date is not None and row_date > target_date:
            return row
    return ws.max_row + 1


def choose_style_source_row(ws, insert_row: int, inserted_count: int) -> int:
    if insert_row > 2:
        return insert_row - 1
    fallback_row = insert_row + inserted_count
    if fallback_row <= ws.max_row:
        return fallback_row
    return 2


def sort_sheet_rows_by_date(ws, date_col: int = 1, header_row: int = 1) -> None:
    data_rows = []
    undated_rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        row_date = excel_date_to_date(values[date_col - 1])
        if row_date is None:
            undated_rows.append(values)
        else:
            data_rows.append((row_date, len(data_rows), values))

    sorted_rows = [values for _, _, values in sorted(data_rows, key=lambda item: (item[0], item[1]))]
    sorted_rows.extend(undated_rows)

    for row_offset, values in enumerate(sorted_rows, start=header_row + 1):
        for col, value in enumerate(values, start=1):
            ws.cell(row_offset, col).value = value


def find_last_data_row(ws, date_col: int = 1) -> int:
    last_row = 1
    for row in range(2, ws.max_row + 1):
        if excel_date_to_date(ws.cell(row, date_col).value) is not None:
            last_row = row
    return last_row


def get_first_sheet(wb, candidates: list[str]):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


def candidate_month_progress_sheet_names(target_month: Optional[int] = None) -> list[str]:
    names = []
    if target_month is not None:
        names.append(f"{int(target_month)}月平台GMV完成情况")
    names.extend([f"{month}月平台GMV完成情况" for month in range(1, 13)])
    names.extend(["平台GMV完成情况", "平台完成情况"])
    seen = set()
    ordered = []
    for name in names:
        if name not in seen:
            ordered.append(name)
            seen.add(name)
    return ordered


def candidate_month_progress_table_names(target_month: Optional[int] = None) -> list[str]:
    return candidate_month_progress_sheet_names(target_month)


def find_month_progress_sheet(wb, target_date: Optional[date] = None):
    if target_date is not None:
        exact_name = f"{target_date.month}月平台GMV完成情况"
        if exact_name in wb.sheetnames:
            return wb[exact_name]
    sheet = get_first_sheet(wb, candidate_month_progress_sheet_names(target_date.month if target_date else None))
    if sheet is not None:
        return sheet
    for name in wb.sheetnames:
        if re.fullmatch(r"\d{1,2}月平台GMV完成情况", str(name).strip()):
            return wb[name]
    return None


def resolve_month_progress_table_name(table_names: dict, target_date: date) -> str:
    configured = [
        table_names.get("month_progress"),
        table_names.get("april_progress"),
    ]
    candidates = [
        str(name).strip()
        for name in configured
        if str(name or "").strip()
    ]
    if candidates:
        return candidates[0]
    candidates.extend(candidate_month_progress_table_names(target_date.month))
    seen = set()
    for name in candidates:
        if name and name not in seen:
            seen.add(name)
            if name == f"{target_date.month}月平台GMV完成情况":
                return name
    for name in candidates:
        if name and name not in seen:
            seen.add(name)
        if name:
            return name
    return BASE_SYNC_TABLES["april_progress"]


def find_channel_sheet_indices(ws) -> dict[str, int]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    return {str(value).strip(): index for index, value in enumerate(headers, start=1) if value is not None}


def find_previous_channel_row(ws, target_date: date) -> Optional[dict]:
    previous = None
    for row in range(2, ws.max_row + 1):
        row_date = excel_date_to_date(ws.cell(row, 1).value)
        if row_date is None or row_date >= target_date:
            continue
        previous = {
            "date": row_date,
            "gmv": float(ws.cell(row, 2).value or 0),
            "actual": float(ws.cell(row, 3).value or 0),
            "refund": float(ws.cell(row, 4).value or 0),
            "refund_rate": float(ws.cell(row, 5).value or 0),
        }
    return previous


def build_day_metrics(data: dict) -> dict:
    entries = data["entries"]
    total_gmv = round2(sum(item["gmv"] for item in entries))
    total_actual = round2(sum(item["actual"] for item in entries))
    total_refund = round2(sum(item["refund"] for item in entries))
    total_refund_rate = safe_div(total_refund, total_gmv)
    totals_override = data.get("totals_override") or {}
    if totals_override:
        total_gmv = round2(totals_override.get("gmv", total_gmv))
        total_actual = round2(totals_override.get("actual", total_actual))
        total_refund = round2(totals_override.get("refund", total_refund))
        total_refund_rate = totals_override.get("refund_rate", safe_div(total_refund, total_gmv))
    platform_set = {item["platform"] for item in entries if item["status"] != "待补齐"}
    full_platforms = {item["platform"] for item in entries if item["status"] in {"完整", "较完整"}}
    top_entry = max(entries, key=lambda item: item["gmv"])
    return {
        "date": data["date"],
        "entries": entries,
        "coverage_note": data["coverage_note"],
        "conclusion": data["conclusion"],
        "total_gmv": total_gmv,
        "total_actual": total_actual,
        "total_refund": total_refund,
        "total_refund_rate": total_refund_rate,
        "platform_count": len(platform_set),
        "full_count": len(full_platforms),
        "top_entry": top_entry,
        "compare_override": data.get("compare_override") or {},
        "month_progress_override": data.get("month_progress_override") or {},
        "youzan_month_override": data.get("youzan_month_override") or {},
    }


def build_split_row(metrics: dict) -> dict:
    row = {value: 0.0 for value in SPLIT_COLUMN_MAP.values()}
    for item in metrics["entries"]:
        split_col = SPLIT_COLUMN_MAP.get((item["platform"], item["account"]))
        if split_col:
            row[split_col] = item["gmv"]
    return row


def build_platform_share_rows(metrics: dict) -> list[tuple[str, float]]:
    rows = []
    for item in metrics["entries"]:
        label = SPLIT_COLUMN_MAP.get((item["platform"], item["account"]), f"{item['platform']}{item['account']}")
        rows.append((label, item["gmv"]))
    return rows


def build_platform_refund_summary_row(metrics: dict) -> list[object]:
    platform_totals = {
        "视频号": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "小红书": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "抖音": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "有赞": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
    }

    for item in metrics["entries"]:
        platform = item["platform"]
        if platform not in platform_totals:
            continue
        platform_totals[platform]["gmv"] += item["gmv"]
        platform_totals[platform]["refund"] += item["refund"]
        platform_totals[platform]["actual"] += item["actual"]

    return [
        metrics["date"],
        round2(platform_totals["视频号"]["gmv"]),
        round2(platform_totals["视频号"]["refund"]),
        round2(platform_totals["视频号"]["actual"]),
        round2(platform_totals["小红书"]["gmv"]),
        round2(platform_totals["小红书"]["refund"]),
        round2(platform_totals["小红书"]["actual"]),
        round2(platform_totals["抖音"]["gmv"]),
        round2(platform_totals["抖音"]["refund"]),
        round2(platform_totals["抖音"]["actual"]),
        round2(platform_totals["有赞"]["gmv"]),
        round2(platform_totals["有赞"]["refund"]),
        round2(platform_totals["有赞"]["actual"]),
    ]


def build_risk_note(metrics: dict) -> str:
    notes: list[str] = []
    entries = metrics["entries"]

    high_refund = [item for item in entries if item["refund_rate"] >= 0.15]
    if high_refund:
        high_refund.sort(key=lambda item: item["refund_rate"], reverse=True)
        notes.append("高退款渠道：" + "、".join(f"{item['platform']}{item['account']} {format_pct(item['refund_rate'])}" for item in high_refund))

    video_big = any(item["platform"] == "视频号" and item["account"] == "大号" for item in entries)
    video_small = any(item["platform"] == "视频号" and item["account"] == "小号" for item in entries)
    if video_big and not video_small:
        notes.append("当前仅纳入视频号大号口径")

    dy_live = sum(item["gmv"] for item in entries if item["platform"] == "抖音" and item["account"] == "直播间")
    dy_card = sum(item["gmv"] for item in entries if item["platform"] == "抖音" and item["account"] == "商品卡")
    if dy_card > 0 and dy_live <= 0:
        notes.append("抖音当天直播间为 0，成交主要来自商品卡")

    xhs_live = any(item["platform"] == "小红书" and item["account"] == "直播" for item in entries)
    xhs_card = any(item["platform"] == "小红书" and item["account"] == "商品卡" for item in entries)
    if xhs_live and xhs_card:
        notes.append("小红书当天已拆分直播和商品卡口径")

    if any(item["status"] != "完整" for item in entries):
        notes.append("当前为阶段性判断，待全渠道数据补全后再下最终结论")

    if metrics["coverage_note"] and "阶段性" in metrics["coverage_note"]:
        notes.append("当前为阶段性判断，待全渠道数据补全后再下最终结论")

    if not notes:
        return "整体数据平稳，可继续关注主力渠道承接情况"
    deduped: list[str] = []
    for note in notes:
        if note not in deduped:
            deduped.append(note)
    return "；".join(deduped)


def build_month_progress(metrics: dict, workbook_path: Path, config: dict, include_current: bool = False) -> dict:
    override = metrics.get("month_progress_override") or {}
    if override:
        target_progress = override.get("target_progress") or []
        normalized_target_progress = []
        for item in target_progress:
            if isinstance(item, dict):
                normalized_target_progress.append(
                    (
                        str(item.get("label", "")).strip(),
                        float(item.get("actual", item.get("completed", 0)) or 0),
                        float(item.get("progress", 0) or 0),
                    )
                )
        return {
            "total_gmv": round2(override.get("total_gmv", 0) or 0),
            "total_actual": round2(override.get("total_actual", 0) or 0),
            "total_refund": round2(override.get("total_refund", 0) or 0),
            "month_goal": float(override.get("month_goal", config.get("month_goal", 0)) or 0),
            "goal_progress": float(override.get("goal_progress", 0) or 0),
            "goal_over": round2(override.get("goal_over", 0) or 0),
            "target_progress": normalized_target_progress,
        }

    wb = load_workbook(workbook_path, data_only=True)
    detail_ws = wb["每日明细+全平台汇总"]

    month = metrics["date"].month
    year = metrics["date"].year

    month_snapshot = aggregate_month_platform_snapshot(
        detail_ws,
        metrics["date"],
        exclude_date=metrics["date"] if include_current else None,
    )

    if include_current:
        current_snapshot = build_month_platform_snapshot(metrics)
        for name, values in current_snapshot.items():
            month_snapshot[name]["gmv"] += values["gmv"]
            month_snapshot[name]["actual"] += values["actual"]
            month_snapshot[name]["refund"] += values["refund"]

    apply_youzan_month_override(month_snapshot, metrics)

    total_gmv = round2(sum(item["gmv"] for item in month_snapshot.values()))
    total_actual = round2(sum(item["actual"] for item in month_snapshot.values()))
    total_refund = round2(sum(item["refund"] for item in month_snapshot.values()))

    month_goal = float(config.get("month_goal", 0) or 0)
    month_targets = config.get("month_targets", {})
    progress = {
        "total_gmv": total_gmv,
        "total_actual": total_actual,
        "total_refund": total_refund,
        "month_goal": month_goal,
        "goal_progress": safe_div(total_actual, month_goal) if month_goal else None,
        "goal_over": round2(total_actual - month_goal) if month_goal else None,
        "target_progress": [],
    }

    for item in MONTH_PROGRESS_ITEMS:
        snapshot_key = item["snapshot_key"]
        target_key = item["target_key"]
        label = item["summary_label"]
        target_value = float(month_targets.get(target_key, 0) or 0)
        if target_value > 0:
            completed = month_snapshot.get(snapshot_key, {}).get("actual", 0.0)
            progress["target_progress"].append((label, completed, safe_div(completed, target_value)))

    card_target = float(month_targets.get("商品卡", month_targets.get("商品卡合计", 0)) or 0)
    if card_target > 0:
        completed = month_snapshot.get("商品卡", {}).get("actual", 0.0)
        progress["target_progress"].append(("商品卡", completed, safe_div(completed, card_target)))

    return progress


def apply_youzan_month_override(month_snapshot: dict[str, dict], metrics: dict) -> None:
    override = metrics.get("youzan_month_override") or {}
    if not override:
        return
    required = ("gmv", "refund", "actual")
    if any(key not in override for key in required):
        raise ValueError("youzan_month_override 缺少 gmv/refund/actual")
    month_snapshot["有赞"] = {
        "gmv": round2(override["gmv"]),
        "refund": round2(override["refund"]),
        "actual": round2(override["actual"]),
    }


def day_exists_in_workbook(workbook_path: Path, target_date: date) -> bool:
    wb = load_workbook(workbook_path, data_only=True)
    for sheet_name in (
        "每日明细+全平台汇总",
        "渠道汇总",
        "分平台GMV日报",
        "分平台销售占比",
        "全平台GMV日报",
        "平台GMV退款汇总",
    ):
        if sheet_name in wb.sheetnames and workbook_has_date(wb[sheet_name], target_date):
            return True
    return False


def format_base_datetime(value: date) -> str:
    return f"{value.isoformat()} 00:00:00"


def parse_base_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return excel_date_to_date(value)


def run_lark_cli(args: list[str]) -> dict:
    completed = subprocess.run(args, capture_output=True, text=True)
    if completed.returncode != 0:
        stdout = completed.stdout.strip()
        stderr = completed.stderr.strip()
        message_parts = [f"lark-cli failed with exit code {completed.returncode}"]
        if stdout:
            message_parts.append(f"stdout:\n{stdout}")
        if stderr:
            message_parts.append(f"stderr:\n{stderr}")
        raise RuntimeError("\n".join(message_parts))
    stdout = completed.stdout.strip()
    if not stdout:
        stderr = completed.stderr.strip()
        command = " ".join(args)
        message = [f"lark-cli returned empty stdout for command: {command}"]
        if stderr:
            message.append(f"stderr:\n{stderr}")
        raise RuntimeError("\n".join(message))
    try:
        return json.loads(stdout)
    except json.JSONDecodeError as exc:
        command = " ".join(args)
        preview = stdout[:1000]
        raise RuntimeError(
            f"lark-cli returned non-JSON stdout for command: {command}\nstdout:\n{preview}"
        ) from exc


def log(message: str) -> None:
    print(message)


def list_base_fields(base_token: str, table_id: str) -> list[dict]:
    result = run_lark_cli(
        [
            "lark-cli",
            "base",
            "+field-list",
            "--base-token",
            base_token,
            "--table-id",
            table_id,
        ]
    )
    data = result.get("data", {}) or {}
    fields = data.get("items")
    if fields is None:
        fields = data.get("fields")
    if not isinstance(fields, list):
        return []
    return fields


def resolve_field_names(base_token: str, table_id: str) -> set[str]:
    names = set()
    for item in list_base_fields(base_token, table_id):
        name = str(item.get("field_name") or item.get("name") or "").strip()
        if name:
            names.add(name)
    if not names:
        raise RuntimeError(
            f"未获取到 Base 表字段：{table_id}。请检查该表是否存在、当前 lark-cli 身份是否有权限、以及 +field-list 是否能正常返回字段列表。"
        )
    return names


def convert_field_value(field_name: str, value):
    if value is None:
        return None
    if field_name == "日期":
        if isinstance(value, str):
            return value
        if isinstance(value, datetime):
            return format_base_datetime(value.date())
        if isinstance(value, date):
            return format_base_datetime(value)
    if field_name in {"占月总%", "完成度"} and isinstance(value, (int, float)):
        return f"{float(value):.2%}"
    return value


def normalize_fields(payload: dict) -> dict:
    result = {}
    for key, value in payload.items():
        converted = convert_field_value(key, value)
        if converted is not None:
            result[key] = converted
    return result


def filter_payload_by_fields(payload: dict, field_names: set[str], table_label: str) -> dict:
    filtered = {}
    skipped = []
    for key, value in payload.items():
        if key in field_names:
            filtered[key] = value
        else:
            skipped.append(key)
    if skipped:
        log(f"{table_label} 跳过 Base 中不存在的字段：{', '.join(skipped)}")
    return filtered


def unpack_base_record_list(result: dict) -> tuple[list[dict], bool]:
    """Normalize both legacy and current lark-cli +record-list envelopes."""
    data = result.get("data", {}) or {}
    raw_items = data.get("items")
    if isinstance(raw_items, list):
        return raw_items, bool(data.get("has_more"))

    rows = data.get("data")
    record_ids = data.get("record_id_list")
    field_names = data.get("fields")
    if not (
        isinstance(rows, list)
        and isinstance(record_ids, list)
        and isinstance(field_names, list)
    ):
        return [], False

    items = []
    for index, record_id in enumerate(record_ids):
        row = rows[index] if index < len(rows) else []
        fields = dict(zip(field_names, row)) if isinstance(row, list) else {}
        items.append({"record_id": record_id, "fields": fields})
    return items, bool(data.get("has_more"))


def list_all_base_records(base_token: str, table_id: str) -> list[dict]:
    items: list[dict] = []
    offset = 0
    while True:
        result = run_lark_cli(
            [
                "lark-cli",
                "base",
                "+record-list",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--offset",
                str(offset),
                "--limit",
                "200",
                "--format",
                "json",
            ]
        )
        batch, has_more = unpack_base_record_list(result)
        items.extend(batch)
        if not has_more:
            break
        offset += 200
    return items


def delete_all_base_records(base_token: str, table_id: str) -> dict:
    items = list_all_base_records(base_token, table_id)
    deleted = 0
    for item in items:
        record_id = item.get("record_id")
        if not record_id:
            continue
        run_lark_cli(
            [
                "lark-cli",
                "base",
                "+record-delete",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--record-id",
                record_id,
                "--yes",
            ]
        )
        deleted += 1
    return {"deleted": deleted}


def record_matches_date(fields: dict, target_date: date) -> bool:
    return parse_base_date(fields.get("日期")) == target_date


def find_matching_base_record_id(base_token: str, table_id: str, target_date: date, payload: dict) -> Optional[str]:
    offset = 0
    while True:
        result = run_lark_cli(
            [
                "lark-cli",
                "base",
                "+record-list",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--offset",
                str(offset),
                "--limit",
                "200",
                "--format",
                "json",
            ]
        )
        items, has_more = unpack_base_record_list(result)
        for item in items:
            fields = item.get("fields", {})
            if not record_matches_date(fields, target_date):
                continue

            if "一级平台" in payload and fields.get("一级平台") != payload.get("一级平台"):
                continue
            if "账号|层级" in payload and fields.get("账号|层级") != payload.get("账号|层级"):
                continue
            if "平台" in payload and fields.get("平台") != payload.get("平台"):
                continue
            return item.get("record_id")
        if not has_more:
            return None
        offset += 200


def build_base_detail_records(metrics: dict) -> list[dict]:
    records = []
    for item in metrics["entries"]:
        records.append(
            drop_none_values(
                {
                "日期": format_base_datetime(metrics["date"]),
                "一级平台": item["platform"],
                "账号|层级": item["account"],
                "GMV|支付金额": item["gmv"],
                "实际销售额": item["actual"],
                "退款额": item["refund"],
                "退款率": item["refund_rate"],
                "订单数": str(item["orders"]) if item["orders"] is not None else None,
                "买家数": str(item["buyers"]) if item["buyers"] is not None else None,
                "数据完整性": item["status"],
                "来源文件|说明": item["source"],
                "字段 1": item.get("live_actual"),
                "字段 2": item.get("card_actual"),
                }
            )
        )

    records.append(
        drop_none_values(
            {
            "日期": format_base_datetime(metrics["date"]),
            "一级平台": "全平台汇总",
            "账号|层级": "每日汇总",
            "GMV|支付金额": metrics["total_gmv"],
            "实际销售额": metrics["total_actual"],
            "退款额": metrics["total_refund"],
            "退款率": metrics["total_refund_rate"],
            "订单数": None,
            "买家数": None,
            "数据完整性": "完整" if "阶段性" not in metrics["coverage_note"] else "较完整",
            "来源文件|说明": metrics["coverage_note"],
            "字段 1": None,
            "字段 2": None,
            }
        )
    )
    return records


def build_base_channel_record(metrics: dict) -> dict:
    return {
        "日期": format_base_datetime(metrics["date"]),
        "已知GMV合计": metrics["total_gmv"],
        "已知实际销售额合计": metrics["total_actual"],
        "已知退款额合计": metrics["total_refund"],
        "已知退款率": metrics["total_refund_rate"],
        "覆盖渠道数": metrics["platform_count"],
        "完整/较完整渠道数": metrics["full_count"],
        "说明": metrics["coverage_note"],
    }


def build_base_share_records(metrics: dict) -> list[dict]:
    return [{"日期": format_base_datetime(metrics["date"]), "平台": label, "销售额": amount} for label, amount in build_platform_share_rows(metrics)]


def build_month_platform_snapshot(metrics: dict) -> dict[str, dict]:
    snapshot = {
        "小红书直播": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "抖音直播": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "商品卡": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "视频号大号": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "视频号小号": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "有赞": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
    }
    for item in metrics["entries"]:
        key = None
        if item["platform"] == "小红书" and item["account"] == "直播":
            key = "小红书直播"
        elif item["platform"] == "抖音" and item["account"] == "直播间":
            key = "抖音直播"
        elif item["platform"] == "视频号" and item["account"] in {"大号", "大号直播"}:
            key = "视频号大号"
            snapshot["商品卡"]["gmv"] += item.get("card_actual", 0.0)
            snapshot["商品卡"]["actual"] += item.get("card_actual", 0.0)
        elif item["platform"] == "视频号" and item["account"] in {"小号", "小号直播"}:
            key = "视频号小号"
            snapshot["商品卡"]["gmv"] += item.get("card_actual", 0.0)
            snapshot["商品卡"]["actual"] += item.get("card_actual", 0.0)
        elif item["platform"] == "有赞":
            key = "有赞"
        elif (item["platform"], item["account"]) in {("小红书", "商品卡"), ("抖音", "商品卡")}:
            key = "商品卡"
        if key:
            snapshot[key]["gmv"] += item["gmv"]
            snapshot[key]["refund"] += item["refund"]
            snapshot[key]["actual"] += item["actual"]
    return snapshot


def aggregate_month_platform_snapshot(detail_ws, target_date: date, exclude_date: Optional[date] = None) -> dict[str, dict]:
    snapshot = {
        "小红书直播": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "抖音直播": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "商品卡": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "视频号大号": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "视频号小号": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
        "有赞": {"gmv": 0.0, "refund": 0.0, "actual": 0.0},
    }
    for row in range(2, detail_ws.max_row + 1):
        row_date = excel_date_to_date(detail_ws.cell(row, 1).value)
        if (
            not row_date
            or row_date.year != target_date.year
            or row_date.month != target_date.month
            or row_date > target_date
        ):
            continue
        if exclude_date is not None and row_date == exclude_date:
            continue
        platform = str(detail_ws.cell(row, 2).value or "").strip()
        account = str(detail_ws.cell(row, 3).value or "").strip()
        key = None
        if platform == "小红书" and account == "直播":
            key = "小红书直播"
        elif platform == "抖音" and account == "直播间":
            key = "抖音直播"
        elif platform == "视频号" and account in {"大号", "大号直播"}:
            key = "视频号大号"
            card_actual = float(detail_ws.cell(row, 13).value or 0)
            snapshot["商品卡"]["gmv"] += card_actual
            snapshot["商品卡"]["actual"] += card_actual
        elif platform == "视频号" and account in {"小号", "小号直播"}:
            key = "视频号小号"
            card_actual = float(detail_ws.cell(row, 13).value or 0)
            snapshot["商品卡"]["gmv"] += card_actual
            snapshot["商品卡"]["actual"] += card_actual
        elif platform == "有赞" and account == "商城":
            key = "有赞"
        elif (platform, account) in {("小红书", "商品卡"), ("抖音", "商品卡")}:
            key = "商品卡"
        if not key:
            continue
        snapshot[key]["gmv"] += float(detail_ws.cell(row, 4).value or 0)
        snapshot[key]["actual"] += float(detail_ws.cell(row, 5).value or 0)
        snapshot[key]["refund"] += float(detail_ws.cell(row, 6).value or 0)
    return snapshot


def aggregate_month_overall_totals(detail_ws, target_date: date, exclude_date: Optional[date] = None) -> dict[str, float]:
    totals = {"gmv": 0.0, "actual": 0.0, "refund": 0.0}
    for row in range(2, detail_ws.max_row + 1):
        row_date = excel_date_to_date(detail_ws.cell(row, 1).value)
        if (
            not row_date
            or row_date.year != target_date.year
            or row_date.month != target_date.month
            or row_date > target_date
        ):
            continue
        if exclude_date is not None and row_date == exclude_date:
            continue
        platform = str(detail_ws.cell(row, 2).value or "").strip()
        if platform != "全平台汇总":
            continue
        totals["gmv"] += float(detail_ws.cell(row, 4).value or 0)
        totals["actual"] += float(detail_ws.cell(row, 5).value or 0)
        totals["refund"] += float(detail_ws.cell(row, 6).value or 0)
    return totals


def build_base_april_progress_records(metrics: dict, config: dict) -> list[dict]:
    progress_override = metrics.get("month_progress_override") or {}
    targets = config.get("month_targets", {}) or {}
    override_items = {
        str(item.get("label", "")).strip(): item
        for item in progress_override.get("target_progress", [])
        if isinstance(item, dict) and str(item.get("label", "")).strip()
    }
    snapshot = build_month_platform_snapshot(metrics)
    records = []
    for label in ("小红书直播", "抖音直播", "商品卡", "视频号大号", "视频号小号", "有赞"):
        item_override = override_items.get(label, {})
        goal = float(item_override.get("goal", targets.get(label if label != "抖音直播" else "抖音直播间", 0)) or 0)
        progress = item_override.get("progress")
        completed = snapshot[label]["gmv"]
        actual = snapshot[label]["actual"]
        refund = snapshot[label]["refund"]
        if item_override.get("completed") is not None and item_override.get("actual") is not None and item_override.get("refund") is not None:
            completed = float(item_override.get("completed") or 0)
            actual = float(item_override.get("actual") or 0)
            refund = float(item_override.get("refund") or 0)
        share = item_override.get("share")
        if share is None and progress_override.get("total_gmv"):
            share = float(completed or 0) / float(progress_override.get("total_gmv") or 1)
        records.append(
            drop_none_values(
                {
                    "平台": label,
                    "成交金额(GMV)": round2(completed or 0),
                    "退款金额": round2(refund or 0),
                    "实际成交额": round2(actual or 0),
                    "月度目标": round2(goal or 0),
                    "占月总%": f"{float(share):.2%}" if share is not None else None,
                    "描述": item_override.get("desc") or "",
                }
            )
        )
    return records


def build_excel_april_progress_rows(
    metrics: dict,
    config: dict,
    month_snapshot: dict[str, dict],
) -> list[list]:
    progress_override = metrics.get("month_progress_override") or {}
    targets = config.get("month_targets", {}) or {}
    override_items = {
        str(item.get("label", "")).strip(): item
        for item in progress_override.get("target_progress", [])
        if isinstance(item, dict) and str(item.get("label", "")).strip()
    }
    rows = []
    for item in MONTH_PROGRESS_ITEMS:
        label = item["sheet_label"]
        summary_label = item["summary_label"]
        target_key = item["target_key"]
        snapshot_key = item["snapshot_key"]
        item_override = override_items.get(summary_label, {})
        goal = float(item_override.get("goal", targets.get(target_key, 0)) or 0)
        completed = round2(item_override.get("completed", month_snapshot[snapshot_key]["gmv"]) or 0)
        actual = round2(item_override.get("actual", month_snapshot[snapshot_key]["actual"]) or 0)
        refund = round2(item_override.get("refund", month_snapshot[snapshot_key]["refund"]) or 0)
        progress = float(item_override.get("progress", safe_div(actual, goal) if goal else 0) or 0)
        remark = f"{item['remark_label']}累计至{metrics['date'].strftime('%-m/%-d')}"
        rows.append(
            [
                label,
                round2(completed or 0),
                round2(refund or 0),
                round2(actual or 0),
                round2(goal or 0),
                float(progress or 0),
                0,
                remark,
                None,
            ]
        )

    total_gmv = round2(progress_override.get("total_gmv", sum(item["gmv"] for item in month_snapshot.values())) or 0)
    total_refund = round2(progress_override.get("total_refund", sum(item["refund"] for item in month_snapshot.values())) or 0)
    total_actual = round2(progress_override.get("total_actual", sum(item["actual"] for item in month_snapshot.values())) or 0)
    month_goal = round2(progress_override.get("month_goal", config.get("month_goal", 0)) or 0)
    goal_progress = float(progress_override.get("goal_progress", safe_div(total_actual, month_goal) if month_goal else 0) or 0)
    total_row = [
        "月度合计",
        total_gmv,
        total_refund,
        total_actual,
        month_goal,
        goal_progress,
        "100.00%",
        "完成度按实际成交额/月度目标计算；当前为阶段性累计" if "阶段性" in metrics["coverage_note"] else "完成度按实际成交额/月度目标计算。",
        None,
    ]
    rows.append(total_row)
    return rows


def build_base_month_progress_records(metrics: dict, config: dict, workbook_path: Path, include_current: bool = False) -> list[dict]:
    wb = load_workbook(workbook_path, data_only=True)
    detail_ws = wb["每日明细+全平台汇总"]
    month_snapshot = aggregate_month_platform_snapshot(
        detail_ws,
        metrics["date"],
        exclude_date=metrics["date"] if include_current else None,
    )
    if include_current:
        current_snapshot = build_month_platform_snapshot(metrics)
        for name, values in current_snapshot.items():
            month_snapshot[name]["gmv"] += values["gmv"]
            month_snapshot[name]["actual"] += values["actual"]
            month_snapshot[name]["refund"] += values["refund"]

    apply_youzan_month_override(month_snapshot, metrics)

    rows = build_excel_april_progress_rows(metrics, config, month_snapshot)
    records = []
    for row in rows:
        records.append(
            drop_none_values(
                {
                    "平台": row[0],
                    "成交金额(GMV)": row[1],
                    "退款金额": row[2],
                    "实际成交额": row[3],
                    "月度目标": row[4],
                    "完成度": row[5],
                    "占月总%": row[6],
                    "描述": row[7],
                    "描述（1）": row[8],
                }
            )
        )
    return records


def build_base_gmv_report_record(metrics: dict, previous: Optional[dict]) -> dict:
    compare_override = metrics.get("compare_override") or {}
    if compare_override:
        gmv_change = float(compare_override.get("gmv_change", 0) or 0)
        actual_change = float(compare_override.get("actual_change", 0) or 0)
        refund_change = float(compare_override.get("refund_change", 0) or 0)
    elif previous and previous["gmv"] > 0 and previous["actual"] > 0 and previous["refund"] > 0:
        gmv_change = (metrics["total_gmv"] - previous["gmv"]) / previous["gmv"]
        actual_change = (metrics["total_actual"] - previous["actual"]) / previous["actual"]
        refund_change = (metrics["total_refund"] - previous["refund"]) / previous["refund"]
    else:
        gmv_change = None
        actual_change = None
        refund_change = None

    top_platform = f"{metrics['top_entry']['platform']}{metrics['top_entry']['account']}"
    return drop_none_values(
        {
            "日期": format_base_datetime(metrics["date"]),
            "直播+私域GMV+商品卡": metrics["total_gmv"],
            "实际销售额": metrics["total_actual"],
            "退款额": metrics["total_refund"],
            "退款率": metrics["total_refund_rate"],
            "较前日GMV": gmv_change,
            "较前日实际": actual_change,
            "较前日退款": refund_change,
            "主力渠道": top_platform,
            "提示": metrics["coverage_note"],
        }
    )


def sync_base_records(base_token: str, table_id: str, records: list[dict], target_date: date, field_names: Optional[set[str]] = None, table_label: Optional[str] = None) -> dict:
    synced = []
    updated = 0
    created = 0
    for payload in records:
        normalized_payload = normalize_fields(payload)
        if field_names is not None:
            normalized_payload = filter_payload_by_fields(normalized_payload, field_names, table_label or table_id)
        if not normalized_payload:
            continue
        existing_record_id = find_matching_base_record_id(base_token, table_id, target_date, normalized_payload)
        if existing_record_id:
            result = run_lark_cli(
                [
                    "lark-cli",
                    "base",
                    "+record-upsert",
                    "--base-token",
                    base_token,
                    "--table-id",
                    table_id,
                    "--record-id",
                    existing_record_id,
                    "--json",
                    json.dumps(normalized_payload, ensure_ascii=False),
                ]
            )
            synced.append(result.get("data", {}))
            updated += 1
            continue

        result = run_lark_cli(
            [
                "lark-cli",
                "base",
                "+record-upsert",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--json",
                json.dumps(normalized_payload, ensure_ascii=False),
            ]
        )
        synced.append(result.get("data", {}))
        created += 1
    return {"updated": updated, "created": created, "items": synced}


def create_base_records(base_token: str, table_id: str, records: list[dict], field_names: Optional[set[str]] = None, table_label: Optional[str] = None) -> dict:
    created = 0
    items = []
    for payload in records:
        normalized_payload = normalize_fields(payload)
        if field_names is not None:
            normalized_payload = filter_payload_by_fields(normalized_payload, field_names, table_label or table_id)
        if not normalized_payload:
            continue
        result = run_lark_cli(
            [
                "lark-cli",
                "base",
                "+record-upsert",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--json",
                json.dumps(normalized_payload, ensure_ascii=False),
            ]
        )
        items.append(result.get("data", {}))
        created += 1
    return {"created": created, "items": items}


def iter_sheet_dict_rows(ws, headers: list[str], start_row: int = 2) -> list[dict]:
    rows: list[dict] = []
    for row in range(start_row, ws.max_row + 1):
        row_dict = {}
        has_value = False
        for col, header in enumerate(headers, start=1):
            if not header or header.startswith("字段"):
                continue
            value = ws.cell(row, col).value
            if value is not None:
                has_value = True
            row_dict[header] = value
        if has_value:
            rows.append(row_dict)
    return rows


def workbook_rows_to_base_records(workbook_path: Path, target_date: Optional[date] = None) -> dict[str, list[dict]]:
    wb = load_workbook(workbook_path, data_only=True)

    detail_records: list[dict] = []
    detail_ws = wb["每日明细+全平台汇总"]
    detail_headers = [str(detail_ws.cell(1, col).value).strip() for col in range(1, detail_ws.max_column + 1) if detail_ws.cell(1, col).value]
    for row in iter_sheet_dict_rows(detail_ws, detail_headers):
        if not row.get("日期"):
            continue
        detail_records.append(
            drop_none_values(
                {
                    "日期": format_base_datetime(excel_date_to_date(row["日期"])),
                    "一级平台": row.get("一级平台"),
                    "账号|层级": row.get("账号|层级"),
                    "GMV|支付金额": row.get("GMV|支付金额"),
                    "实际销售额": row.get("实际销售额"),
                    "退款额": row.get("退款额"),
                    "退款率": row.get("退款率"),
                    "订单数": str(row.get("订单数")) if row.get("订单数") is not None else None,
                    "买家数": str(row.get("买家数")) if row.get("买家数") is not None else None,
                    "数据完整性": row.get("数据完整性"),
                    "来源文件|说明": row.get("来源文件|说明"),
                }
            )
        )

    channel_records: list[dict] = []
    channel_ws = wb["渠道汇总"]
    channel_headers = [str(channel_ws.cell(1, col).value).strip() for col in range(1, channel_ws.max_column + 1) if channel_ws.cell(1, col).value]
    for row in iter_sheet_dict_rows(channel_ws, channel_headers):
        if not row.get("日期"):
            continue
        channel_records.append(
            drop_none_values(
                {
                    "日期": format_base_datetime(excel_date_to_date(row["日期"])),
                    "已知GMV合计": row.get("已知GMV合计"),
                    "已知实际销售额合计": row.get("已知实际销售额合计"),
                    "已知退款额合计": row.get("已知退款额合计"),
                    "已知退款率": row.get("已知退款率"),
                    "覆盖渠道数": row.get("覆盖渠道数"),
                    "完整/较完整渠道数": row.get("完整/较完整渠道数"),
                    "说明": row.get("说明"),
                }
            )
        )

    share_records: list[dict] = []
    share_ws = wb["分平台销售占比"]
    share_headers = [str(share_ws.cell(1, col).value).strip() for col in range(1, share_ws.max_column + 1) if share_ws.cell(1, col).value]
    for row in iter_sheet_dict_rows(share_ws, share_headers):
        if not row.get("日期") or not row.get("平台"):
            continue
        share_records.append(
            drop_none_values(
                {
                    "日期": format_base_datetime(excel_date_to_date(row["日期"])),
                    "平台": row.get("平台"),
                    "销售额": row.get("销售额"),
                }
            )
        )

    april_records: list[dict] = []
    april_ws = find_month_progress_sheet(wb, target_date)
    if april_ws is not None:
        for row in range(2, 9):
            label = april_ws.cell(row, 1).value
            if label is None:
                continue
            share_value = april_ws.cell(row, 7).value
            if isinstance(share_value, (int, float)):
                share_value = f"{float(share_value):.2%}"
            april_records.append(
                drop_none_values(
                    {
                        "平台": label,
                        "成交金额(GMV)": april_ws.cell(row, 2).value,
                        "退款金额": april_ws.cell(row, 3).value,
                        "实际成交额": april_ws.cell(row, 4).value,
                        "月度目标": april_ws.cell(row, 5).value,
                        "占月总%": share_value,
                        "描述": april_ws.cell(row, 8).value,
                        "描述（1）": april_ws.cell(row, 9).value,
                    }
                )
            )

    gmv_report_records: list[dict] = []
    gmv_ws = wb["全平台GMV日报"]
    gmv_headers = [str(gmv_ws.cell(1, col).value).strip() for col in range(1, gmv_ws.max_column + 1) if gmv_ws.cell(1, col).value]
    for row in iter_sheet_dict_rows(gmv_ws, gmv_headers):
        if not row.get("日期"):
            continue
        record = {}
        for key in ("日期", "直播+私域GMV+商品卡", "实际销售额", "退款额", "退款率", "较前日GMV", "较前日实际", "较前日退款", "主力渠道", "提示"):
            value = row.get(key)
            if key == "日期" and value is not None:
                value = format_base_datetime(excel_date_to_date(value))
            if value is not None:
                record[key] = value
        gmv_report_records.append(record)

    return {
        "detail": detail_records,
        "channel": channel_records,
        "share": share_records,
        "april_progress": april_records,
        "gmv_report": gmv_report_records,
    }


def replace_base_table(base_token: str, table_id: str, records: list[dict], field_names: Optional[set[str]] = None, table_label: Optional[str] = None) -> dict:
    delete_result = delete_all_base_records(base_token, table_id)
    create_result = create_base_records(base_token, table_id, records, field_names, table_label)
    return {
        "deleted": delete_result["deleted"],
        "created": create_result["created"],
        "items": create_result["items"],
    }


def sync_to_base(metrics: dict, config: dict, workbook_path: Path) -> dict:
    base_token = str(config.get("base_token", "")).strip()
    if not base_token:
        raise ValueError("配置文件缺少 base_token，无法同步飞书 Base")

    table_names = config.get("base_tables", {})
    detail_table = str(table_names.get("detail", BASE_SYNC_TABLES["detail"]))
    channel_table = str(table_names.get("channel", BASE_SYNC_TABLES["channel"]))
    share_table = str(table_names.get("share", BASE_SYNC_TABLES["share"]))
    april_progress_table = resolve_month_progress_table_name(table_names, metrics["date"])
    gmv_report_table = str(table_names.get("gmv_report", BASE_SYNC_TABLES["gmv_report"]))
    detail_fields = resolve_field_names(base_token, detail_table)
    channel_fields = resolve_field_names(base_token, channel_table)
    share_fields = resolve_field_names(base_token, share_table)
    april_progress_fields = resolve_field_names(base_token, april_progress_table)
    gmv_report_fields = resolve_field_names(base_token, gmv_report_table)
    previous = None
    if workbook_path.exists():
        previous = find_previous_channel_row(load_workbook(workbook_path, data_only=True)["渠道汇总"], metrics["date"])

    workbook_records = workbook_rows_to_base_records(workbook_path, metrics["date"])
    include_current_in_month_progress = True
    month_progress_records = build_base_month_progress_records(
        metrics,
        config,
        workbook_path,
        include_current=include_current_in_month_progress,
    )
    detail_result = sync_base_records(base_token, detail_table, build_base_detail_records(metrics), metrics["date"], detail_fields, "每日明细+全平台汇总")
    channel_result = sync_base_records(base_token, channel_table, [build_base_channel_record(metrics)], metrics["date"], channel_fields, "渠道汇总")
    share_result = sync_base_records(base_token, share_table, build_base_share_records(metrics), metrics["date"], share_fields, "分平台销售占比")
    april_progress_result = replace_base_table(base_token, april_progress_table, month_progress_records, april_progress_fields, april_progress_table)
    gmv_report_result = sync_base_records(base_token, gmv_report_table, [build_base_gmv_report_record(metrics, previous)], metrics["date"], gmv_report_fields, "全平台GMV日报")
    return {
        "detail": detail_result,
        "channel": channel_result,
        "share": share_result,
        "april_progress": april_progress_result,
        "gmv_report": gmv_report_result,
    }


def build_summary_text(metrics: dict, previous: Optional[dict], month_progress: Optional[dict]) -> tuple[str, list[str], list[str], str]:
    total_gmv = format_wan(metrics["total_gmv"])
    total_actual = format_wan(metrics["total_actual"])
    total_refund = format_wan(metrics["total_refund"])
    refund_rate = format_pct(metrics["total_refund_rate"])

    delta_text = ""
    compare_override = metrics.get("compare_override") or {}
    if compare_override:
        compare_date = str(compare_override.get("date", "")).strip()
        gmv_delta = float(compare_override.get("gmv_change", 0) or 0)
        actual_delta = float(compare_override.get("actual_change", 0) or 0)
        refund_delta = float(compare_override.get("refund_change", 0) or 0)
        gmv_word = "增长" if gmv_delta >= 0 else "下降"
        actual_word = "增长" if actual_delta >= 0 else "下降"
        refund_word = "增长" if refund_delta >= 0 else "下降"
        delta_text = (
            f"较{compare_date}，"
            f"GMV{gmv_word} {abs(gmv_delta):.2%}，"
            f"实际销售额{actual_word} {abs(actual_delta):.2%}，"
            f"退款额{refund_word} {abs(refund_delta):.2%}。"
        )
    elif previous and previous["gmv"] > 0 and previous["actual"] > 0 and previous["refund"] > 0:
        gmv_delta = (metrics["total_gmv"] - previous["gmv"]) / previous["gmv"]
        actual_delta = (metrics["total_actual"] - previous["actual"]) / previous["actual"]
        refund_delta = (metrics["total_refund"] - previous["refund"]) / previous["refund"]
        gmv_word = "增长" if gmv_delta >= 0 else "下降"
        actual_word = "增长" if actual_delta >= 0 else "下降"
        refund_word = "增长" if refund_delta >= 0 else "下降"
        delta_text = (
            f"较{previous['date'].month}月{previous['date'].day}日，"
            f"GMV{gmv_word} {abs(gmv_delta):.2%}，"
            f"实际销售额{actual_word} {abs(actual_delta):.2%}，"
            f"退款额{refund_word} {abs(refund_delta):.2%}。"
        )

    coverage_note = metrics["coverage_note"] or "口径完整。"
    summary = (
        f"截至{metrics['date'].month}月{metrics['date'].day}日，全平台 GMV {total_gmv}，"
        f"实际销售额 {total_actual}，退款额 {total_refund}，退款率 {refund_rate}。"
        f"{delta_text}{coverage_note}"
    )

    platform_lines = []
    grouped: dict[str, list[dict]] = {}
    for item in metrics["entries"]:
        grouped.setdefault(item["platform"], []).append(item)

    for platform in ("小红书", "抖音", "有赞", "视频号"):
        items = grouped.get(platform, [])
        if not items:
            continue
        platform_gmv = sum(item["gmv"] for item in items)
        share = safe_div(platform_gmv, metrics["total_gmv"])
        if platform == "小红书":
            live = next((item for item in items if item["account"] == "直播"), None)
            card = next((item for item in items if item["account"] == "商品卡"), None)
            if live and card:
                platform_lines.append(
                    f"小红书合计 GMV {format_wan(platform_gmv)}，其中直播 {format_wan(live['gmv'])}、商品卡 {format_wan(card['gmv'])}，占全平台 {format_pct(share)}；直播退款率 {format_pct(live['refund_rate'])}，商品卡退款率 {format_pct(card['refund_rate'])}。"
                )
            else:
                platform_lines.append(f"小红书 GMV {format_wan(platform_gmv)}，占全平台 {format_pct(share)}。")
        elif platform == "抖音":
            live = next((item for item in items if item["account"] == "直播间"), None)
            card = next((item for item in items if item["account"] == "商品卡"), None)
            if live and card:
                platform_lines.append(
                    f"抖音合计 GMV {format_wan(platform_gmv)}，其中直播间 {format_wan(live['gmv'])}、非直播 {format_wan(card['gmv'])}，占全平台 {format_pct(share)}；直播间退款率 {format_pct(live['refund_rate'])}，非直播退款率 {format_pct(card['refund_rate'])}。"
                )
            else:
                platform_lines.append(f"抖音 GMV {format_wan(platform_gmv)}，占全平台 {format_pct(share)}。")
        elif platform == "有赞":
            item = items[0]
            platform_lines.append(f"有赞 GMV {format_wan(item['gmv'])}，退款率 {format_pct(item['refund_rate'])}。")
        elif platform == "视频号":
            if len(items) == 2:
                big = next((item for item in items if item["account"] == "大号"), None)
                small = next((item for item in items if item["account"] == "小号"), None)
                parts = []
                if big:
                    parts.append(f"大号 {format_wan(big['gmv'])}")
                if small:
                    parts.append(f"小号 {format_wan(small['gmv'])}")
                platform_lines.append(f"视频号合计 GMV {format_wan(platform_gmv)}，其中" + "、".join(parts) + f"，占全平台 {format_pct(share)}。")
            else:
                item = items[0]
                platform_lines.append(f"视频号{item['account']} GMV {format_wan(item['gmv'])}，退款率 {format_pct(item['refund_rate'])}。")

    month_lines: list[str] = []
    if month_progress and month_progress["month_goal"]:
        goal_over_text = (
            f"已超目标 {format_wan(month_progress['goal_over'])}"
            if month_progress["goal_over"] >= 0
            else f"距目标还差 {format_wan(abs(month_progress['goal_over']))}"
        )
        month_lines.append(
            f"截至{metrics['date'].month}月{metrics['date'].day}日，累计 GMV {format_wan(month_progress['total_gmv'])}，累计退款 {format_wan(month_progress['total_refund'])}，累计实际销售额 {format_wan(month_progress['total_actual'])}，对照月度总目标 {format_wan(month_progress['month_goal'])}，当前完成度 {format_pct(month_progress['goal_progress'])}，{goal_over_text}。"
        )
        if month_progress["target_progress"]:
            target_text = "，".join(
                f"{label}实际销售额 {format_wan(actual)}，完成 {format_pct(progress)}"
                for label, actual, progress in month_progress["target_progress"]
            )
            month_lines.append(target_text + "。")

    conclusion = metrics["conclusion"] or build_risk_note(metrics)
    return summary, platform_lines, month_lines, conclusion


def build_card(title: str, summary: str, platform_lines: list[str], month_lines: list[str], conclusion: str, metrics: dict) -> dict:
    elements = [
        {"tag": "div", "text": {"tag": "lark_md", "content": f"**核心数据**\n{summary}"}},
        {
            "tag": "div",
            "fields": [
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**GMV**\n{format_wan(metrics['total_gmv'])}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**实际销售额**\n{format_wan(metrics['total_actual'])}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**退款额**\n{format_wan(metrics['total_refund'])}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**退款率**\n{format_pct(metrics['total_refund_rate'])}"}},
            ],
        },
        {"tag": "hr"},
        {"tag": "div", "text": {"tag": "lark_md", "content": "**分平台表现**\n" + "\n".join(f"- {line}" for line in platform_lines)}},
    ]
    if metrics.get("coverage_note"):
        elements.extend([
            {"tag": "hr"},
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**口径说明**\n{metrics['coverage_note']}"}},
        ])
    if month_lines:
        elements.extend([
            {"tag": "hr"},
            {"tag": "div", "text": {"tag": "lark_md", "content": "**月度进度**\n" + "\n".join(f"- {line}" for line in month_lines)}},
        ])
    elements.extend([
        {"tag": "hr"},
        {"tag": "div", "text": {"tag": "lark_md", "content": f"**总结判断**\n{conclusion}"}},
    ])
    return {
        "msg_type": "interactive",
        "card": {
            "config": {"wide_screen_mode": True},
            "header": {"template": "blue", "title": {"tag": "plain_text", "content": title}},
            "elements": elements,
        },
    }


def build_markdown(title: str, summary: str, platform_lines: list[str], month_lines: list[str], conclusion: str, coverage_note: str) -> str:
    sections = [
        f"# {title}",
        "",
        "## 核心数据",
        summary,
        "",
        "## 分平台表现",
        *[f"- {line}" for line in platform_lines],
    ]
    if coverage_note:
        sections.extend(["", "## 口径说明", coverage_note])
    if month_lines:
        sections.extend(["", "## 月度进度", *[f"- {line}" for line in month_lines]])
    sections.extend(["", "## 总结判断", conclusion, ""])
    return "\n".join(sections)


def post_webhook(webhook: str, payload: dict) -> dict:
    req = urllib.request.Request(
        webhook,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode("utf-8"))


def backup_workbook(workbook_path: Path, target_date: date) -> Path:
    backup_dir = workbook_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{workbook_path.stem}_{target_date.isoformat()}_{timestamp}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def update_workbook(workbook_path: Path, metrics: dict, coverage_note: str, config: dict) -> None:
    wb = load_workbook(workbook_path)
    detail_ws = wb["每日明细+全平台汇总"]
    channel_ws = wb["渠道汇总"]
    split_ws = get_first_sheet(wb, ["分平台GMV日报"])
    platform_share_ws = get_first_sheet(wb, ["平台销售占比长表", "分平台销售占比"])
    refund_summary_ws = get_first_sheet(wb, ["平台GMV退款汇总"])
    report_ws = wb["执行总监日报展示版"] if "执行总监日报展示版" in wb.sheetnames else None
    note_ws = wb["数据说明"] if "数据说明" in wb.sheetnames else None
    april_progress_ws = find_month_progress_sheet(wb, metrics["date"])
    gmv_report_ws = get_first_sheet(wb, ["全平台GMV日报", "全平台GMV报"])

    for ws in (detail_ws, channel_ws, split_ws, platform_share_ws, gmv_report_ws, refund_summary_ws, report_ws):
        if ws is not None:
            delete_rows_by_date(ws, metrics["date"])

    detail_rows_to_add = len(metrics["entries"]) + 1
    start_row = find_insert_row_by_date(detail_ws, metrics["date"])
    detail_ws.insert_rows(start_row, detail_rows_to_add)
    detail_template_row = choose_style_source_row(detail_ws, start_row, detail_rows_to_add)
    for offset, item in enumerate(metrics["entries"]):
        row_num = start_row + offset
        copy_row_style(detail_ws, detail_template_row, row_num, detail_ws.max_column)
        values = [
            metrics["date"],
            item["platform"],
            item["account"],
            item["gmv"],
            item["actual"],
            item["refund"],
            item["refund_rate"],
            item["orders"],
            item["buyers"],
            item["status"],
            item["source"],
            item.get("live_actual"),
            item.get("card_actual"),
        ]
        for col, value in enumerate(values, start=1):
            detail_ws.cell(row_num, col).value = value

    total_row_num = start_row + len(metrics["entries"])
    copy_row_style(detail_ws, detail_template_row, total_row_num, detail_ws.max_column)
    total_values = [
        metrics["date"],
        "全平台汇总",
        "每日汇总",
        metrics["total_gmv"],
        metrics["total_actual"],
        metrics["total_refund"],
        metrics["total_refund_rate"],
        None,
        None,
        "完整" if "阶段性" not in coverage_note else "较完整",
        coverage_note,
    ]
    for col, value in enumerate(total_values, start=1):
        detail_ws.cell(total_row_num, col).value = value

    channel_template_row = max(2, find_last_data_row(channel_ws))
    channel_row = channel_template_row + 1
    copy_row_style(channel_ws, channel_template_row, channel_row, channel_ws.max_column)
    channel_values = [
        metrics["date"],
        metrics["total_gmv"],
        metrics["total_actual"],
        metrics["total_refund"],
        metrics["total_refund_rate"],
        metrics["platform_count"],
        metrics["full_count"],
        coverage_note,
    ]
    for col, value in enumerate(channel_values, start=1):
        channel_ws.cell(channel_row, col).value = value

    if split_ws:
        split_template_row = max(2, find_last_data_row(split_ws))
        split_row = split_template_row + 1
        copy_row_style(split_ws, split_template_row, split_row, split_ws.max_column)
        split_headers = {str(split_ws.cell(1, col).value).strip(): col for col in range(1, split_ws.max_column + 1) if split_ws.cell(1, col).value}
        split_values = build_split_row(metrics)
        split_ws.cell(split_row, 1).value = metrics["date"]
        for name, value in split_values.items():
            col = split_headers.get(name)
            if col:
                split_ws.cell(split_row, col).value = value

    if platform_share_ws:
        share_rows = build_platform_share_rows(metrics)
        share_start_row = find_insert_row_by_date(platform_share_ws, metrics["date"])
        platform_share_ws.insert_rows(share_start_row, len(share_rows))
        share_template_row = choose_style_source_row(platform_share_ws, share_start_row, len(share_rows))
        for offset, (label, amount) in enumerate(share_rows):
            row_num = share_start_row + offset
            copy_row_style(platform_share_ws, share_template_row, row_num, platform_share_ws.max_column)
            platform_share_ws.cell(row_num, 1).value = metrics["date"]
            platform_share_ws.cell(row_num, 2).value = label
            platform_share_ws.cell(row_num, 3).value = amount

    if gmv_report_ws:
        previous = find_previous_channel_row(channel_ws, metrics["date"])
        gmv_template_row = max(2, find_last_data_row(gmv_report_ws))
        report_row = gmv_template_row + 1
        copy_row_style(gmv_report_ws, gmv_template_row, report_row, gmv_report_ws.max_column)
        record = build_base_gmv_report_record(metrics, previous)
        gmv_headers = {str(gmv_report_ws.cell(1, col).value).strip(): col for col in range(1, gmv_report_ws.max_column + 1) if gmv_report_ws.cell(1, col).value}
        for field_name, value in record.items():
            col = gmv_headers.get(field_name)
            if col:
                gmv_report_ws.cell(report_row, col).value = metrics["date"] if field_name == "日期" else value

    if refund_summary_ws:
        refund_template_row = max(2, find_last_data_row(refund_summary_ws))
        refund_row = refund_template_row + 1
        copy_row_style(refund_summary_ws, refund_template_row, refund_row, refund_summary_ws.max_column)
        refund_values = build_platform_refund_summary_row(metrics)
        for col, value in enumerate(refund_values, start=1):
            refund_summary_ws.cell(refund_row, col).value = value

    if report_ws:
        previous = find_previous_channel_row(channel_ws, metrics["date"])
        report_ws.insert_rows(REPORT_INSERT_ROW, 1)
        copy_row_style(report_ws, REPORT_INSERT_ROW - 1, REPORT_INSERT_ROW, report_ws.max_column)
        prev_gmv = previous["gmv"] if previous and previous["gmv"] else None
        prev_actual = previous["actual"] if previous and previous["actual"] else None
        prev_refund = previous["refund"] if previous and previous["refund"] else None
        top_platform = f"{metrics['top_entry']['platform']}{metrics['top_entry']['account']}"
        report_values = [
            metrics["date"],
            metrics["total_gmv"],
            metrics["total_actual"],
            metrics["total_refund"],
            metrics["total_refund_rate"],
            safe_div(metrics["total_gmv"] - prev_gmv, prev_gmv) if prev_gmv else None,
            safe_div(metrics["total_actual"] - prev_actual, prev_actual) if prev_actual else None,
            safe_div(metrics["total_refund"] - prev_refund, prev_refund) if prev_refund else None,
            top_platform,
            coverage_note,
        ]
        for col, value in enumerate(report_values, start=1):
            report_ws.cell(REPORT_INSERT_ROW, col).value = value
        report_ws["B3"] = f"{metrics['date'].strftime('%Y-%m-01')} 至 {metrics['date'].isoformat()}"

    if note_ws:
        note_text = f"已更新到 {metrics['date'].isoformat()}（经营数据）；{coverage_note}"
        note_ws["B2"] = note_text
        note_ws["B3"] = note_text
        note_ws["B4"] = note_text
        note_ws["C4"] = note_text
        note_ws["B5"] = "退款率口径：全平台汇总退款率按当日各平台退款额合计 / GMV合计 计算。"
        note_ws["C5"] = note_ws["B5"].value

    if april_progress_ws:
        month_snapshot = aggregate_month_platform_snapshot(detail_ws, metrics["date"])
        apply_youzan_month_override(month_snapshot, metrics)
        april_rows = build_excel_april_progress_rows(metrics, config, month_snapshot)
        start_row = 2
        for offset, row_values in enumerate(april_rows):
            row_num = start_row + offset
            for col, value in enumerate(row_values, start=1):
                april_progress_ws.cell(row_num, col).value = value
            april_progress_ws.cell(row_num, 6).number_format = "0.00%"
            if row_num < start_row + len(april_rows) - 1 and isinstance(april_progress_ws.cell(row_num, 7).value, (int, float)):
                april_progress_ws.cell(row_num, 7).number_format = "0.00%"
        april_progress_ws.cell(9, 1).value = (
            f"说明：完成度按实际成交额/月度目标计算。数据已更新至{metrics['date'].strftime('%-m月%-d日')}；"
            f"{coverage_note}；月度目标合计{round2(config.get('month_goal', 0) / 10000)}万。"
        )
        for col in range(2, april_progress_ws.max_column + 1):
            april_progress_ws.cell(9, col).value = None

    for ws in (detail_ws, channel_ws, split_ws, platform_share_ws, gmv_report_ws, refund_summary_ws):
        if ws is not None:
            sort_sheet_rows_by_date(ws)

    wb.save(workbook_path)


def write_outputs(output_dir: Path, target_date: date, markdown: str, card: dict) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    markdown_path = output_dir / f"report_{target_date.strftime('%Y_%m_%d')}.md"
    payload_path = output_dir / f"card_{target_date.strftime('%Y_%m_%d')}.json"
    markdown_path.write_text(markdown, encoding="utf-8")
    save_json(payload_path, card)
    return markdown_path, payload_path


def main() -> None:
    parser = argparse.ArgumentParser(description="执行总监日报一键流水线：更新总表、生成飞书卡片、可选发送")
    parser.add_argument("--input", required=True, help="标准化输入 JSON 路径")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="配置文件路径")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="执行总监总表路径")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="日报输出目录")
    parser.add_argument("--sync-base", action="store_true", help="同步到飞书 Base")
    parser.add_argument("--send", action="store_true", help="通过飞书 webhook 正式发送")
    parser.add_argument("--send-only", action="store_true", help="只发送飞书，不再写 Excel")
    parser.add_argument("--skip-workbook", action="store_true", help="只生成日报内容，不写 Excel")
    parser.add_argument("--dry-run", action="store_true", help="只检查和输出结果，不发送飞书")
    args = parser.parse_args()

    input_path = Path(args.input)
    config = load_config(Path(args.config))
    data = validate_input(load_json(input_path))
    metrics = build_day_metrics(data)
    workbook_path = Path(args.workbook)

    previous = None
    if workbook_path.exists():
        previous = find_previous_channel_row(load_workbook(workbook_path, data_only=True)["渠道汇总"], metrics["date"])

    if args.send_only:
        args.send = True
        args.skip_workbook = True

    write_workbook = not args.skip_workbook and not args.dry_run

    if write_workbook:
        backup_path = backup_workbook(workbook_path, metrics["date"])
        update_workbook(workbook_path, metrics, data["coverage_note"], config)
    else:
        backup_path = None

    include_current_in_month_progress = not write_workbook
    month_progress = (
        build_month_progress(metrics, workbook_path, config, include_current=include_current_in_month_progress)
        if workbook_path.exists()
        else None
    )
    summary, platform_lines, month_lines, conclusion = build_summary_text(metrics, previous, month_progress)
    title = f"{metrics['date'].year}年{metrics['date'].month}月{metrics['date'].day}日执行总监经营日报"
    card = build_card(title, summary, platform_lines, month_lines, conclusion, metrics)
    markdown = build_markdown(title, summary, platform_lines, month_lines, conclusion, metrics["coverage_note"])
    markdown_path, payload_path = write_outputs(Path(args.output_dir), metrics["date"], markdown, card)

    result = {
        "date": metrics["date"].isoformat(),
        "workbook": str(workbook_path),
        "backup": str(backup_path) if backup_path else None,
        "markdown": str(markdown_path),
        "payload": str(payload_path),
        "summary": summary,
    }

    if args.sync_base and not args.dry_run:
        result["base_sync"] = sync_to_base(metrics, config, workbook_path)

    if args.dry_run or not args.send:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        print(json.dumps(card, ensure_ascii=False, indent=2))
        return

    webhook = str(config.get("webhook") or WEBHOOK_FALLBACK).strip()
    if not webhook.startswith("https://"):
        raise ValueError("配置文件中的 webhook 不合法")
    send_result = post_webhook(webhook, card)
    result["send_result"] = send_result
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"流水线执行失败：{exc}", file=sys.stderr)
        sys.exit(1)
